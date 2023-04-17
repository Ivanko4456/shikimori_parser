import os.path
import sys
import tkinter as tk
import openpyxl
import requests
from bs4 import BeautifulSoup


def getHTML(url: str) -> str:
    headers = {
        'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/107.0'
                      '.0.0 Safari/537.36 OPR/93.0.0.0',
        'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,'
                  'application/signed-exchange;v=b3;q=0.9'}
    
    html = requests.get(url=url, headers=headers)
    if html.status_code == 200:
        return html.text
    else:
        tk.messagebox.showerror(title='Ошибка', message='Пользователь не найден')


def get_anime_list(animehtml: str):
    anime_names, nums, categs, comments, values, seriesCurr, seriesMax = None, None, None, None, None, None, None
    soup = BeautifulSoup(animehtml, "html.parser")
    try:
        rawlist = soup.find_all('span', class_="name-ru")
        anime_names = [name.text for name in rawlist]

        raw_comments = soup.find_all('div', class_='rate-text')
        comments = [comment.text for comment in raw_comments]

        rawnums = soup.find_all('td', class_='index')
        nums = [int(num.text) for num in rawnums]

        raw_vals = soup.find_all('span', attrs={'data-field': 'score'})
        values = [value.text for value in raw_vals]

        raw_sers = soup.find_all('span', attrs={'data-field': 'episodes'})
        seriesMax = [serie['data-max'] for serie in raw_sers]
        seriesCurr = [serie.text for serie in raw_sers]

        rawcategories = soup.find('div', class_='list-groups').find_all('div', class_='subheadline m5')
        categs = [categ.text for categ in rawcategories]
    except AttributeError:
        tk.messagebox.showerror(title='Ошибка', message='Похоже, у пользователя нет списка')
    return anime_names, nums, categs, comments, values, (seriesCurr, seriesMax)


def make_tuple(names: list, nums: list, categs: list, comments: list, values: list, series: tuple) -> list:
    categ = -1
    anime = []
    for i in range(len(nums)):
        if nums[i] != 1:
            anime.append((categs[categ], nums[i], names[i], comments[i], values[i], series[0][i], series[1][i]))
        elif nums[i] == 1:
            # if not isfirst:
            #     anime.append(('', '', '', '', '', '', ''))
            categ += 1
            anime.append((categs[categ], nums[i], names[i], comments[i], values[i], series[0][i], series[1][i]))
    return anime


# def csv_write(anime: list, username: str, new: bool) -> None:
#     if not new:
#         with open(f"{username}'s list 1.csv", 'w') as file:
#             names = ['Статус', 'Название', 'Эпизоды', 'Оценка', 'Комментарий']
#             file_writer = csv.DictWriter(file, delimiter=";", lineterminator="\r", fieldnames=names)
#             file_writer.writeheader()
#             for a in anime:
#                 a = list(a)
#                 print(a)
#                 if a[4] == '–' or a[4] == '':
#                     a[4] = ' '
#                 file_writer.writerow({'Статус': a[0], 'Название': a[2], 'Оценка': f'{a[4]} | 10' if a[4] != ' ' else '',
#                                       'Комментарий': a[3],
#                                       'Эпизоды': f'{a[5]} | {a[6]}' if a[5] != '' else ''})
#     else:
#         for table in range(1, 100):
#             if not os.path.exists(f"{username}'s list {table}.csv"):
#                 with open(f"{username}'s list {table}.csv", 'w') as file:
#                     names = ['Статус', 'Название', 'Оценка', 'Комментарий']
#                     file_writer = csv.DictWriter(file, delimiter=";", lineterminator="\r", fieldnames=names)
#                     file_writer.writeheader()
#                     for a in anime:
#                         a = list(a)
#
#                         if a[4] == '–' or a[4] == '':
#                             a[4] = ' '
#                         file_writer.writerow({'Статус': a[0], 'Название': a[2], 'Оценка': a[4], 'Комментарий': a[3]})
#                 return


def xlsx_write(anime: list, username: str, new: bool):
    names = sorted([a[2] for a in anime])
    nmlen = sorted(list(map(len, names)), reverse=True)[0]

    statuses = sorted([a[0] for a in anime])
    smlen = sorted(list(map(len, statuses)), reverse=True)[0]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.column_dimensions['A'].width = smlen + 3
    ws.column_dimensions['B'].width = nmlen + 3
    ws.column_dimensions['E'].width = 50
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 8
    ws['A1'], ws['B1'], ws['C1'], ws['D1'], ws['E1'] = 'Статус', 'Название', 'Эпизоды', 'Оценка', 'Комментарий'
    i = 2
    for a in anime:
        a = list(a)
        if a[4] == '–' or a[4] == '':
            a[4] = ' '
         
        ws[f'A{i}'] = a[0]  # status
        
        ws[f'B{i}'] = a[2]  # name
        
        ws[f'C{i}'] = f'{a[5]} / {a[6]}' if a[5] != '' else ''  # series
        
        ws[f'D{i}'] = int(a[4]) if a[4] != ' ' else ''  # mark
        ws[f'D{i}'].alignment = openpyxl.styles.Alignment(horizontal="left")
        
        ws[f'E{i}'] = a[3]  #comment
        
        i += 1
    path = f"{username}'s list 1.xlsx"
    if new:
        for table in range(1, 1000):
            if not os.path.exists(f"{username}'s list {table}.xlsx"):
                path = f"{username}'s list {table}.xlsx"
                break
    wb.save(path)


def resource_path(relative_path):
    base_path = getattr(sys, '_MEIPASS', os.path.dirname(os.path.abspath(__file__)))
    return os.path.join(base_path, relative_path)


def main(user, new=False):

    url = f'https://shikimori.one/{user}/list/anime'

    html = getHTML(url)
    names, nums, cats, comments, values, series = get_anime_list(html)
    anime = make_tuple(names, nums, cats, comments, values, series)
    xlsx_write(anime=anime, username=user, new=new)
    tk.messagebox.showinfo(title='Готово', message='Таблица успешно создана!')
    exit()
